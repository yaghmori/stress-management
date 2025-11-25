"""
Excel export service for generating reports with RTL and Persian support.
"""

from typing import List, Dict, Any
from datetime import date
from pathlib import Path
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# RTL text support
try:
    import arabic_reshaper
    import bidi.algorithm as bidi_algorithm
    RTL_SUPPORT = True
except ImportError:
    RTL_SUPPORT = False

from app.config.translation_manager import TranslationManager
from app.config.date_utils import format_date_for_display


logger = logging.getLogger(__name__)


class ExcelService:
    """Service for Excel report generation with RTL support."""
    
    def __init__(self, translation_manager: TranslationManager) -> None:
        """
        Initialize Excel service.
        
        Args:
            translation_manager: Translation manager instance
        """
        self.t = translation_manager.t
        self._init_rtl_support()
    
    def _init_rtl_support(self) -> None:
        """Initialize RTL text support."""
        self.has_rtl_support = RTL_SUPPORT
        if not RTL_SUPPORT:
            logger.warning("RTL support libraries not available. Persian text may not render correctly.")
    
    def _prepare_rtl_text(self, text: str, reshape: bool = False) -> str:
        """
        Prepare text for RTL rendering in Excel.
        
        Args:
            text: Input text (may contain Persian/Arabic characters)
            reshape: Whether to reshape the text (default: False, Excel handles it natively)
            
        Returns:
            Text for RTL rendering (reshaped only if requested)
        """
        if not text:
            return text
        
        # For Excel, we typically don't reshape as Excel handles RTL natively
        # Reshaping can cause character placement issues with some words
        # Only reshape if explicitly requested (for data cells that might need it)
        if reshape and self.has_rtl_support:
            try:
                # Check if text contains Persian/Arabic characters
                has_rtl = any('\u0600' <= char <= '\u06FF' or '\u0750' <= char <= '\u077F' 
                             for char in text)
                
                if has_rtl:
                    # Only reshape if explicitly requested
                    reshaped_text = arabic_reshaper.reshape(text)
                    return reshaped_text
            except Exception as e:
                logger.warning(f"Error processing RTL text: {e}")
        
        # Return text as-is for headers and most content
        # Excel's native RTL support handles Persian text correctly
        return text
    
    def export_stress_report(self, file_path: Path, user: Dict[str, Any],
                            logs: List[Dict[str, Any]], 
                            start_date: date, end_date: date) -> bool:
        """
        Export stress logs to Excel report with RTL support.
        
        Args:
            file_path: Path to save Excel file
            user: User data dictionary
            logs: List of stress log entries
            start_date: Start date of report period
            end_date: End date of report period
            
        Returns:
            True if export successful, False otherwise
        """
        if not EXCEL_SUPPORT:
            logger.error("openpyxl library not available")
            return False
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = self.t("reports")  # Use text as-is for sheet title
            
            # Define styles - Excel will handle font fallback automatically
            # Use Vazir for Persian text, Excel will fallback to Arial if not available
            font_name = 'Vazir'  # Excel will use system default if font not found
            
            title_font = Font(name=font_name, size=18, bold=True, color='FFFFFF')
            title_fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
            
            heading_font = Font(name=font_name, size=14, bold=True, color='FFFFFF')
            heading_fill = PatternFill(start_color='34495e', end_color='34495e', fill_type='solid')
            
            header_font = Font(name=font_name, size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='34495e', end_color='34495e', fill_type='solid')
            
            label_font = Font(name=font_name, size=10, bold=True)
            label_fill = PatternFill(start_color='ecf0f1', end_color='ecf0f1', fill_type='solid')
            
            summary_font = Font(name=font_name, size=11, bold=True, color='FFFFFF')
            summary_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
            
            data_font = Font(name=font_name, size=10)
            number_font = Font(name=font_name, size=10, bold=True, color='1a1a1a')  # Dark color for numbers
            
            # RTL alignment
            rtl_alignment = Alignment(horizontal='right', vertical='center', text_rotation=0)
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # Border
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            row = 1
            
            # Title - don't reshape, Excel handles it natively
            ws.merge_cells(f'A{row}:B{row}')
            title_cell = ws[f'A{row}']
            title_cell.value = self.t("reports")  # Use text as-is without reshaping
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = center_alignment
            row += 2
            
            # User information - RTL layout (value in column A, label in column B)
            # Don't reshape labels, Excel handles Persian text natively
            user_info = [
                [self._prepare_rtl_text(user.get('username', 'N/A'), reshape=True), self.t("username") + ":"],
                [self._prepare_rtl_text(format_date_for_display(start_date), reshape=True), self.t("reports_date_from") + ":"],
                [self._prepare_rtl_text(format_date_for_display(end_date), reshape=True), self.t("reports_date_to") + ":"],
                [str(len(logs)), "تعداد رکوردها:"]
            ]
            
            for info_row in user_info:
                value_cell = ws[f'A{row}']
                label_cell = ws[f'B{row}']
                
                value_cell.value = info_row[0]
                value_cell.font = data_font
                value_cell.alignment = rtl_alignment
                value_cell.border = thin_border
                
                label_cell.value = info_row[1]
                label_cell.font = label_font
                label_cell.fill = label_fill
                label_cell.alignment = rtl_alignment
                label_cell.border = thin_border
                row += 1
            
            row += 1
            
            # Summary statistics
            if logs:
                avg_stress = sum(log.get('stress_level', 0) for log in logs) / len(logs)
                total_sleep = sum(log.get('sleep_hours', 0) or 0 for log in logs)
                avg_sleep = total_sleep / len(logs) if logs else 0
                total_activity = sum(log.get('physical_activity', 0) or 0 for log in logs)
                avg_activity = total_activity / len(logs) if logs else 0
                
                # Summary data - RTL order (value in column A, label in column B)
                # Don't reshape labels, Excel handles Persian text natively
                summary_data = [
                    [f"{avg_stress:.2f}/10", self.t("stress_level") + " " + self.t("weekly_average") + ":"],
                    [f"{avg_sleep:.2f}", self.t("sleep_hours") + " " + self.t("weekly_average") + ":"],
                    [f"{avg_activity:.2f} " + self.t("minutes"), self.t("physical_activity") + " " + self.t("weekly_average") + ":"]
                ]
                
                for summary_row in summary_data:
                    value_cell = ws[f'A{row}']
                    label_cell = ws[f'B{row}']
                    
                    value_cell.value = summary_row[0]
                    value_cell.font = number_font  # Use bold dark font for numbers
                    value_cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    value_cell.alignment = rtl_alignment
                    value_cell.border = thin_border
                    
                    label_cell.value = summary_row[1]
                    label_cell.font = summary_font
                    label_cell.fill = summary_fill
                    label_cell.alignment = rtl_alignment
                    label_cell.border = thin_border
                    row += 1
                
                row += 1
            
            # Stress logs table
            if logs:
                # Table heading - don't reshape, Excel handles it natively
                ws.merge_cells(f'A{row}:E{row}')
                table_heading = ws[f'A{row}']
                table_heading.value = self.t("stress_history")  # Use text as-is without reshaping
                table_heading.font = heading_font
                table_heading.fill = heading_fill
                table_heading.alignment = center_alignment
                row += 1
                
                # Table headers - RTL order (rightmost column first)
                # Don't reshape headers, Excel handles Persian text natively
                headers = [
                    self.t("notes"),
                    self.t("physical_activity"),
                    self.t("sleep_hours"),
                    self.t("stress_level"),
                    self.t("date")
                ]
                
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment  # Center headers for better readability
                    cell.border = thin_border
                
                row += 1
                
                # Table data - RTL order
                for log in logs:
                    log_date = log.get('date', '')
                    if isinstance(log_date, str):
                        date_str = format_date_for_display(log_date)
                    else:
                        date_str = format_date_for_display(log_date)
                    
                    stress_level = log.get('stress_level', '')
                    sleep_hours = log.get('sleep_hours', '') if log.get('sleep_hours') else '-'
                    physical_activity = log.get('physical_activity', '') if log.get('physical_activity') else '-'
                    notes = str(log.get('notes', '')) if log.get('notes') else '-'
                    
                    # RTL order: notes, physical_activity, sleep_hours, stress_level, date
                    # Reshape data content (notes, dates) but not headers
                    row_data = [
                        self._prepare_rtl_text(notes, reshape=True),
                        physical_activity if physical_activity != '-' else '-',
                        sleep_hours if sleep_hours != '-' else '-',
                        stress_level,
                        self._prepare_rtl_text(date_str, reshape=True)
                    ]
                    
                    for col_idx, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = value
                        
                        # Use number font for numeric values
                        if isinstance(value, (int, float)) or (isinstance(value, str) and value.replace('.', '').replace('-', '').isdigit()):
                            cell.font = number_font
                        else:
                            cell.font = data_font
                        
                        # Alternating row colors
                        if row % 2 == 0:
                            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
                        
                        cell.alignment = rtl_alignment
                        cell.border = thin_border
                    
                    row += 1
            
            # Set column widths
            ws.column_dimensions['A'].width = 25  # Value/Notes column
            ws.column_dimensions['B'].width = 18  # Physical activity column
            ws.column_dimensions['C'].width = 15  # Sleep hours column
            ws.column_dimensions['D'].width = 12  # Stress level column
            ws.column_dimensions['E'].width = 15  # Date column
            
            # Set sheet direction to RTL
            ws.sheet_view.rightToLeft = True
            
            # Save workbook
            wb.save(str(file_path))
            logger.info(f"Excel report exported successfully to: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting Excel report: {e}", exc_info=True)
            return False

